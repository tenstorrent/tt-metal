# SPDX-FileCopyrightText: © 2026 Tenstorrent AI ULC
# SPDX-License-Identifier: Apache-2.0
"""
Reduce a full-inference tracy CSV to ONE layer of measured ops per sub-block.

Builds on split_inference_perf_to_xlsx.py: same per-sub-block sheet structure
(Speaker_Encoder, Talker_Prefill, Talker_Decode, CP_Prefill, CP_Decode), but
each multi-layer decoder block is truncated to layer 0 only — so kernel-time
totals are directly comparable across blocks.

Layer-0 detection: each decoder layer begins with a *pre-attention* LayerNorm
acting on the residual stream (IN0 X == hidden_size). The 2nd such layernorm
marks the start of layer 1, so layer 0 is everything before that mark.

Usage:
    python split_inference_perf_one_layer.py <ops_perf_results.csv> [--out perf.xlsx]
"""
from __future__ import annotations

import argparse
import csv
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# Re-use the phase classifier from the full-block splitter so we don't drift.
sys.path.insert(0, str(Path(__file__).parent))
from split_inference_perf_to_xlsx import DURATION_COL, KEEP_COLS, _logical, _ts, classify, detect_phases  # noqa: E402

# Hidden-size markers identifying a "pre-attention layernorm" (= layer entry).
HIDDEN_TALKER = 2048
HIDDEN_CP = 1024


def _is_layer_entry_norm(row: dict, hidden: int) -> bool:
    """A LayerNorm whose IN0 X-dim equals the block's hidden_size — the residual
    stream norm at the top of each decoder layer. Excludes per-head q/k norms
    (X=head_dim=128) and the final norm (also X=hidden but that comes after the
    last layer; we cap at num_layers-1 boundaries so it doesn't matter)."""
    if not row["OP CODE"].startswith("LayerNorm"):
        return False
    x = _logical(row.get("INPUT_0_X_PAD[LOGICAL]", ""))
    return x == hidden


def truncate_to_layer0(rows: list[dict], hidden: int) -> list[dict]:
    """Return ops belonging to the first decoder layer only.

    Each decoder layer has TWO LayerNorms at IN0 X = hidden (pre-attn + post-attn).
    Layer 0 starts at the 1st such layernorm and ends just before the 3rd
    (= start of layer 1). Anything before the 1st (e.g. pre-layer Untilize of
    input embeddings) is excluded.
    """
    rows_sorted = sorted(rows, key=_ts)
    seen = 0
    start = None
    cut = len(rows_sorted)
    for i, r in enumerate(rows_sorted):
        if _is_layer_entry_norm(r, hidden):
            seen += 1
            if seen == 1:
                start = i
            elif seen == 3:
                cut = i
                break
    if start is None:
        return []
    return rows_sorted[start:cut]


def truncate_decode_trace_to_layer0(rows: list[dict], hidden: int) -> list[dict]:
    """Per-trace truncation: each captured CP_Decode trace contains 5 layers; we
    want the first layer of just one trace. Group rows by METAL TRACE ID, take
    the first trace's layer-0 slice.
    """
    by_trace: dict[str, list[dict]] = defaultdict(list)
    for r in rows:
        by_trace[(r.get("METAL TRACE ID", "") or "").strip()].append(r)
    if not by_trace:
        return []
    # Pick the trace whose first row has the smallest HOST START TS.
    trace_ids = sorted(by_trace.keys(), key=lambda t: _ts(min(by_trace[t], key=_ts)))
    first_trace_rows = by_trace[trace_ids[0]]
    return truncate_to_layer0(first_trace_rows, hidden)


SHEET_ORDER = [
    ("speaker_encoder", "Speaker_Encoder", None, None),  # leave as-is
    ("talker_prefill", "Talker_Prefill", HIDDEN_TALKER, "single"),
    ("talker_decode", "Talker_Decode", HIDDEN_TALKER, "single"),
    ("cp_prefill", "CP_Prefill", HIDDEN_CP, "single"),
    ("cp_decode", "CP_Decode", HIDDEN_CP, "per_trace"),
]


def write_xlsx(blocks: dict[str, list[dict]], out_path: Path, fieldnames: list[str], phases: dict[str, int]) -> None:
    keep = [c for c in KEEP_COLS if c in fieldnames]
    wb = openpyxl.Workbook()
    summary = wb.active
    summary.title = "Summary"
    summary.append(["Block (1-layer view)", "Ops", "Total kernel us", "Top-5 ops by us"])
    for cell in summary[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")

    for key, title, hidden, mode in SHEET_ORDER:
        rows = blocks.get(key, [])
        if hidden is not None:
            if mode == "per_trace":
                rows = truncate_decode_trace_to_layer0(rows, hidden)
            else:
                rows = truncate_to_layer0(rows, hidden)
        ws = wb.create_sheet(title)
        ws.append(keep)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
        rows_sorted = sorted(rows, key=_ts)
        for r in rows_sorted:
            ws.append([r.get(c, "") for c in keep])
        for i, col in enumerate(keep, start=1):
            ws.column_dimensions[get_column_letter(i)].width = max(10, min(40, len(col) + 2))
        if rows_sorted and DURATION_COL in keep:
            dur_col_idx = keep.index(DURATION_COL) + 1
            col_letter = get_column_letter(dur_col_idx)
            data_range = f"{col_letter}2:{col_letter}{len(rows_sorted) + 1}"
            ws.conditional_formatting.add(
                data_range, DataBarRule(start_type="min", end_type="max", color="63BE7B", showValue=True)
            )
        ws.freeze_panes = "A2"

        total_us = sum(int(r.get(DURATION_COL) or 0) for r in rows_sorted) / 1000.0
        op_us: dict[str, float] = defaultdict(float)
        for r in rows_sorted:
            op_us[r.get("OP CODE", "")] += int(r.get(DURATION_COL) or 0) / 1000.0
        top = sorted(op_us.items(), key=lambda x: -x[1])[:5]
        top_str = ", ".join(f"{n.replace('DeviceOperation', '')}={d:.1f}" for n, d in top)
        summary.append([title, len(rows_sorted), round(total_us, 2), top_str])

    summary.append([])
    summary.append(["Phase boundary (HOST START TS, ns)", "Value"])
    for k, v in phases.items():
        summary.append([f"  {k}", v])
    summary.column_dimensions["A"].width = 32
    summary.column_dimensions["B"].width = 8
    summary.column_dimensions["C"].width = 16
    summary.column_dimensions["D"].width = 100
    summary.freeze_panes = "A2"

    wb.save(out_path)
    print(f"Wrote {out_path}")


def main():
    parser = argparse.ArgumentParser(
        description="Split full-inference tracy CSV into multi-sheet xlsx; truncate decoder blocks to layer 0."
    )
    parser.add_argument("csv", help="Path to ops_perf_results_*.csv from tracy")
    parser.add_argument("--out", default=None, help="Output xlsx path (default: <csv_stem>.layer0.xlsx)")
    args = parser.parse_args()
    csv_path = Path(args.csv)
    if not csv_path.exists():
        print(f"ERROR: {csv_path} does not exist", file=sys.stderr)
        sys.exit(1)
    out_path = Path(args.out) if args.out else csv_path.with_suffix(".layer0.xlsx")

    print(f"Reading: {csv_path}")
    with open(csv_path, newline="") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        fieldnames = reader.fieldnames or []
    print(f"  Read {len(rows)} rows.")

    phases = detect_phases(rows)
    print("Phase boundaries (ns):")
    for k, v in phases.items():
        print(f"  {k:>22}: {v}")

    blocks = classify(rows, phases)
    print("\nFull-block sizes (before layer-0 truncation):")
    for k, t, *_ in SHEET_ORDER:
        print(f"  {k:18s}: {len(blocks.get(k, []))}")

    write_xlsx(blocks, out_path, fieldnames, phases)


if __name__ == "__main__":
    main()
