# SPDX-FileCopyrightText: © 2026 Tenstorrent AI ULC
# SPDX-License-Identifier: Apache-2.0
"""
Split a full-inference tracy CSV into per-sub-block sheets in an xlsx.

Generates ONE measurement-iteration worth of ops per block (no warmup, no
trace capture, no weight uploads). Phase detection uses HOST START TS plus
shape markers from the demo flow.

Sheets produced (one per inference sub-block):
    Summary
    Speaker_Encoder      — pre-Talker-prefill non-traced ops (audio encoder + ICL)
    Talker_Prefill       — non-traced ops with X=2048 hidden, seq>=2
    Talker_Decode        — captured trace, one frame's worth of ops
    CP_Prefill           — captured trace, one frame's worth (seq=2)
    CP_Decode            — 14 captured traces (codebooks 2-15), one frame's worth each
    CP_Decode_Trace_##   — (optional) one sheet per CP decode trace_id, with --per-trace

Phase detection (HOST START TS):
    T_compute_start       — first non-weight-tilize compute op
    T_talker_prefill_start — first Matmul with X=2048 and Y>=2 (logical)
    T_warmup_start        — first Matmul with X=2048 and Y=1   (decoder warmup begins)

Anything between T_warmup_start and the trace captures, plus all untraced ops
that tracy reorders after the trace captures, are dropped (these are warmup
runs + inter-trace H2D, not steady-state inference).

For traced blocks we keep METAL TRACE REPLAY SESSION ID == 1 only — that's
exactly one frame's worth of replay. The audio decoder is pure-PyTorch in this
demo, so no Audio_Decoder sheet is generated.

Usage:
    python split_inference_perf_to_xlsx.py <ops_perf_results.csv> [--out perf.xlsx]
"""
from __future__ import annotations

import argparse
import csv
import sys
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

KEEP_COLS = [
    "OP CODE",
    "CORE COUNT",
    "MATH FIDELITY",
    "DEVICE KERNEL DURATION [ns]",
    "INPUT_0_Y_PAD[LOGICAL]",
    "INPUT_0_X_PAD[LOGICAL]",
    "INPUT_0_LAYOUT",
    "INPUT_0_DATATYPE",
    "INPUT_0_MEMORY",
    "INPUT_1_Y_PAD[LOGICAL]",
    "INPUT_1_X_PAD[LOGICAL]",
    "OUTPUT_0_Y_PAD[LOGICAL]",
    "OUTPUT_0_X_PAD[LOGICAL]",
    "OUTPUT_0_MEMORY",
    "METAL TRACE ID",
    "METAL TRACE REPLAY SESSION ID",
]

DURATION_COL = "DEVICE KERNEL DURATION [ns]"
TILIZE_OPS = {
    "TilizeDeviceOperation",
    "UntilizeDeviceOperation",
    "UntilizeWithUnpaddingDeviceOperation",
    "TilizeWithValPaddingDeviceOperation",
}


def _logical(s: str) -> int | None:
    s = (s or "").strip()
    if not s:
        return None
    if "[" in s and "]" in s:
        try:
            return int(s[s.index("[") + 1 : s.index("]")])
        except ValueError:
            return None
    try:
        return int(s)
    except ValueError:
        return None


def _y(r: dict) -> int | None:
    return _logical(r.get("INPUT_0_Y_PAD[LOGICAL]", ""))


def _x(r: dict) -> int | None:
    return _logical(r.get("INPUT_0_X_PAD[LOGICAL]", ""))


def _ts(r: dict) -> int:
    return int(r.get("HOST START TS") or 0)


def _trace_id(r: dict) -> str:
    return (r.get("METAL TRACE ID", "") or "").strip()


def _session_id(r: dict) -> str:
    return (r.get("METAL TRACE REPLAY SESSION ID", "") or "").strip()


def detect_phases(rows: list[dict]) -> dict[str, int]:
    """Return phase boundary timestamps."""
    untraced_sorted = sorted([r for r in rows if not _trace_id(r)], key=_ts)
    traced_sess1 = [r for r in rows if _trace_id(r) and _session_id(r) == "1"]

    T_compute_start = next(
        (_ts(r) for r in untraced_sorted if r["OP CODE"] not in TILIZE_OPS),
        0,
    )
    T_talker_prefill_start = next(
        (_ts(r) for r in untraced_sorted if r["OP CODE"].startswith("Matmul") and _x(r) == 2048 and (_y(r) or 0) >= 2),
        0,
    )
    T_warmup_start = next(
        (_ts(r) for r in untraced_sorted if r["OP CODE"].startswith("Matmul") and _x(r) == 2048 and _y(r) == 1),
        0,
    )
    T_gen_start = min((_ts(r) for r in traced_sess1), default=0)
    T_gen_end = max((_ts(r) for r in rows if _trace_id(r)), default=0)

    return {
        "compute_start": T_compute_start,
        "talker_prefill_start": T_talker_prefill_start,
        "warmup_start": T_warmup_start,
        "gen_start": T_gen_start,
        "gen_end": T_gen_end,
    }


def classify(rows: list[dict], phases: dict[str, int]) -> dict[str, list[dict]]:
    """Bucket every row into a named block (or 'drop').

    Traced rows are routed by per-TRACE majority vote on IN0 shape (Talker uses
    hidden=2048, CP uses hidden=1024; CP prefill has Y_logical>=2, CP decode
    has Y_logical=1). This avoids putting the few non-hidden-shaped ops of a
    trace (e.g. attention-internal QK/AV matmuls at K=128) into the wrong
    block.
    """
    blocks: dict[str, list[dict]] = defaultdict(list)

    cs = phases["compute_start"]
    tps = phases["talker_prefill_start"]
    ws = phases["warmup_start"]

    # Step 1: vote per traced trace_id using only its session=1 rows.
    sess1_by_trace: dict[str, list[dict]] = defaultdict(list)
    for r in rows:
        if _trace_id(r) and _session_id(r) == "1":
            sess1_by_trace[_trace_id(r)].append(r)
    trace_label: dict[str, str] = {}
    for tid, sub in sess1_by_trace.items():
        votes: Counter = Counter()
        for r in sub:
            if _x(r) == 2048:
                votes["talker_decode"] += 1
            elif _x(r) == 1024:
                votes["cp_decode" if _y(r) == 1 else "cp_prefill"] += 1
        trace_label[tid] = votes.most_common(1)[0][0] if votes else "talker_decode"

    # Step 2: bucket every row.
    for r in rows:
        tid = _trace_id(r)
        ts = _ts(r)

        if tid:
            # Take session=1 only — one frame of replay.
            if _session_id(r) != "1":
                continue
            blocks[trace_label.get(tid, "talker_decode")].append(r)
        else:
            # Only [T_compute_start, T_warmup_start] holds real inference work.
            # Anything before is weight upload; anything from T_warmup_start onward
            # is warmup/capture or trace-reordered noise (audio decoder is pure
            # PyTorch in this demo, so no on-device audio ops exist).
            if ts < cs:
                blocks["weight_upload"].append(r)  # drop
            elif cs <= ts < tps:
                blocks["speaker_encoder"].append(r)  # incl. tiny ICL on-device tail
            elif tps <= ts < ws:
                blocks["talker_prefill"].append(r)
            else:
                blocks["warmup_or_reordered"].append(r)  # drop

    return blocks


# Sheets to actually write (in order). Internal-only buckets like weight_upload,
# warmup_capture, inter_trace are intentionally dropped per user request.
SHEET_ORDER = [
    ("speaker_encoder", "Speaker_Encoder"),
    ("talker_prefill", "Talker_Prefill"),
    ("talker_decode", "Talker_Decode"),
    ("cp_prefill", "CP_Prefill"),
    ("cp_decode", "CP_Decode"),
]


def write_xlsx(
    blocks: dict[str, list[dict]],
    out_path: Path,
    fieldnames: list[str],
    phases: dict[str, int],
    per_trace: bool = False,
) -> None:
    keep = [c for c in KEEP_COLS if c in fieldnames]

    wb = openpyxl.Workbook()
    summary = wb.active
    summary.title = "Summary"
    summary.append(["Block", "Ops", "Total kernel us", "Top-5 ops by us"])
    for cell in summary[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")

    for key, title in SHEET_ORDER:
        rows = blocks.get(key, [])
        ws = wb.create_sheet(title)
        ws.append(keep)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
        # Sort by HOST START TS to keep the natural execution order on the sheet.
        rows_sorted = sorted(rows, key=_ts)
        for r in rows_sorted:
            ws.append([r.get(c, "") for c in keep])
        for i, col in enumerate(keep, start=1):
            ws.column_dimensions[get_column_letter(i)].width = max(10, min(40, len(col) + 2))
        if rows_sorted and DURATION_COL in keep:
            dur_col_idx = keep.index(DURATION_COL) + 1
            col_letter = get_column_letter(dur_col_idx)
            data_range = f"{col_letter}2:{col_letter}{len(rows_sorted) + 1}"
            ws.conditional_formatting.add(
                data_range,
                DataBarRule(start_type="min", end_type="max", color="63BE7B", showValue=True),
            )
        ws.freeze_panes = "A2"

        total_us = sum(int(r.get(DURATION_COL) or 0) for r in rows_sorted) / 1000.0
        op_us: dict[str, float] = defaultdict(float)
        for r in rows_sorted:
            op_us[r.get("OP CODE", "")] += int(r.get(DURATION_COL) or 0) / 1000.0
        top = sorted(op_us.items(), key=lambda x: -x[1])[:5]
        top_str = ", ".join(f"{n.replace('DeviceOperation','')}={d:.1f}" for n, d in top)
        summary.append([title, len(rows_sorted), round(total_us, 2), top_str])

    # Optional: split CP_Decode into per-trace sheets (one per codebook).
    if per_trace:
        cp_rows = sorted(blocks.get("cp_decode", []), key=_ts)
        by_trace: dict[str, list[dict]] = defaultdict(list)
        for r in cp_rows:
            by_trace[_trace_id(r)].append(r)
        # Sort traces by their first row's HOST START TS so sheet order matches replay order.
        sorted_tids = sorted(by_trace.keys(), key=lambda t: _ts(by_trace[t][0]))
        for idx, tid in enumerate(sorted_tids):
            rows_t = by_trace[tid]
            ws = wb.create_sheet(f"CP_Decode_{idx:02d}_tid{tid}")
            ws.append(keep)
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
            for r in rows_t:
                ws.append([r.get(c, "") for c in keep])
            for i, col in enumerate(keep, start=1):
                ws.column_dimensions[get_column_letter(i)].width = max(10, min(40, len(col) + 2))
            if rows_t and DURATION_COL in keep:
                dur_col_idx = keep.index(DURATION_COL) + 1
                col_letter = get_column_letter(dur_col_idx)
                data_range = f"{col_letter}2:{col_letter}{len(rows_t) + 1}"
                ws.conditional_formatting.add(
                    data_range,
                    DataBarRule(start_type="min", end_type="max", color="63BE7B", showValue=True),
                )
            ws.freeze_panes = "A2"

    # Append phase timestamps to summary as a footer.
    summary.append([])
    summary.append(["Phase boundary (HOST START TS, ns)", "Value"])
    for k, v in phases.items():
        summary.append([f"  {k}", v])

    summary.column_dimensions["A"].width = 38
    summary.column_dimensions["B"].width = 16
    summary.column_dimensions["C"].width = 16
    summary.column_dimensions["D"].width = 100
    summary.freeze_panes = "A2"

    wb.save(out_path)
    print(f"Wrote {out_path}")


def main():
    parser = argparse.ArgumentParser(
        description="Split full-inference tracy CSV into multi-sheet xlsx (final-iteration ops only)."
    )
    parser.add_argument("csv", help="Path to ops_perf_results_*.csv from tracy")
    parser.add_argument("--out", default=None, help="Output xlsx path (default: <csv_stem>.blocks.xlsx)")
    parser.add_argument(
        "--per-trace",
        action="store_true",
        help="Add one sheet per CP_Decode trace_id (14 extra sheets, one per codebook).",
    )
    args = parser.parse_args()
    csv_path = Path(args.csv)
    if not csv_path.exists():
        print(f"ERROR: {csv_path} does not exist", file=sys.stderr)
        sys.exit(1)
    out_path = Path(args.out) if args.out else csv_path.with_suffix(".blocks.xlsx")

    print(f"Reading: {csv_path}")
    with open(csv_path, newline="") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        fieldnames = reader.fieldnames or []
    print(f"  Read {len(rows)} rows.")

    phases = detect_phases(rows)
    print("Phase boundaries (ns):")
    for k, v in phases.items():
        print(f"  {k:>22}: {v}")

    blocks = classify(rows, phases)
    print("\nBlock sizes:")
    for k, t in SHEET_ORDER:
        print(f"  {k:18s}: {len(blocks.get(k, []))}")
    dropped = sum(len(blocks.get(k, [])) for k in ("weight_upload", "warmup_or_reordered"))
    print(f"  (dropped: weight_upload + warmup_or_reordered = {dropped})")

    write_xlsx(blocks, out_path, fieldnames, phases, per_trace=args.per_trace)


if __name__ == "__main__":
    main()
