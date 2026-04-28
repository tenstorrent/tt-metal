#!/usr/bin/env python3
"""
Combined Excel/CSV Processor
=============================
This script combines the functionality of:
1. CSV processing with opcodes extraction (perf_analyzer.py)
2. Excel formatting with data bars (format_excel_databars.py)
3. FPS calculations (add_fps_calculation.py)

Usage:
    python excel_processor_combined.py <input_file> [output_directory]

    input_file: Can be either .csv or .xlsx file
    output_directory: Optional, defaults to current directory
"""

import argparse
import sys
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Font, NamedStyle, PatternFill
from openpyxl.utils import get_column_letter


class ExcelProcessor:
    def __init__(self, input_file, output_dir="."):
        self.input_file = Path(input_file)
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        self.workbook = None
        self.excel_file = None

    def process_csv_to_excel(self):
        """Process CSV file and create Excel with opcodes and formatted columns"""
        print(f"📄 Processing CSV file: {self.input_file}")

        # Read CSV
        df = pd.read_csv(self.input_file)

        # Define required columns, with fallback options for different CSV formats
        required_cols = ["OP CODE", "CORE COUNT", "DEVICE KERNEL DURATION [ns]"]

        # Add MATH FIDELITY if available
        if "MATH FIDELITY" in df.columns:
            required_cols.append("MATH FIDELITY")

        optional_cols = []

        # Check which input/output columns exist and add them
        for prefix in ["INPUT_0", "OUTPUT_0"]:
            for suffix in ["W", "Z", "Y", "X"]:
                # Try different column name formats
                possible_names = [
                    f"{prefix}_{suffix}",  # Original format
                    f"{prefix}_{suffix}_PAD[LOGICAL]",  # New format with _PAD[LOGICAL]
                ]
                for col_name in possible_names:
                    if col_name in df.columns:
                        optional_cols.append(col_name)
                        break

            # Add layout, datatype, and memory columns
            for attr in ["LAYOUT", "DATATYPE", "MEMORY"]:
                col_name = f"{prefix}_{attr}"
                if col_name in df.columns:
                    optional_cols.append(col_name)

        # Select available columns
        col_list = required_cols + optional_cols
        available_cols = [col for col in col_list if col in df.columns]

        print(f"📊 Processing {len(available_cols)} available columns")
        df = df[available_cols]

        # Find first non-nan row in CORE COUNT
        first_index = df["CORE COUNT"].first_valid_index()
        if first_index is not None:
            df = df.iloc[first_index:]

        # Add OP ID as first column (row numbers starting from 1)
        df.reset_index(drop=True, inplace=True)
        df.insert(0, "OP ID", range(1, len(df) + 1))

        # Rename columns for brevity
        df.rename(columns=lambda x: x.replace("INPUT_", "IN_"), inplace=True)
        df.rename(columns=lambda x: x.replace("OUTPUT_", "OUT_"), inplace=True)

        # Clean up OP CODE values
        df = df.replace("InterleavedToShardedDeviceOperation", "I2S", regex=True)
        df = df.replace("ShardedToInterleavedDeviceOperation", "S2I", regex=True)
        df = df.replace("DeviceOperation", "", regex=True)
        df = df.replace("DEV_0_", "", regex=True)

        # Convert data types
        df = df.convert_dtypes()

        # Calculate total duration before formatting
        total_duration = df["DEVICE KERNEL DURATION [ns]"].sum()
        print(f"📈 Total duration: {total_duration:,} ns")

        # Generate output filename
        base_name = self.input_file.stem
        self.excel_file = self.output_dir / f"{base_name}_processed_complete.xlsx"

        # Save to Excel
        df.to_excel(self.excel_file, index=False, sheet_name="Performance_Data")
        print(f"✅ Excel file created: {self.excel_file}")

        return self.excel_file

    def load_excel_file(self):
        """Load existing Excel file"""
        print(f"📊 Loading Excel file: {self.input_file}")
        self.excel_file = self.input_file
        return self.excel_file

    def add_formatted_columns_and_databars(self):
        """Add formatted duration column with data bars to all worksheets"""
        print(f"\n🎨 Adding formatted columns and data bars...")

        # Load workbook
        wb = openpyxl.load_workbook(self.excel_file)

        # Create number style with comma separator
        comma_style = NamedStyle(name="comma_no_decimal")
        comma_style.number_format = "#,##0"

        try:
            wb.add_named_style(comma_style)
        except ValueError:
            # Style already exists
            pass

        processed_sheets = 0

        for sheet_idx, ws in enumerate(wb.worksheets, 1):
            sheet_name = ws.title
            print(f"  [{sheet_idx:2d}/{len(wb.worksheets)}] Processing sheet: '{sheet_name}'")

            try:
                # Find the DEVICE KERNEL DURATION column
                duration_col_idx = None
                duration_col_letter = None

                # Check first row for headers
                for col_idx, cell in enumerate(ws[1], 1):
                    if cell.value and "DEVICE KERNEL DURATION [ns]" in str(cell.value):
                        duration_col_idx = col_idx
                        duration_col_letter = get_column_letter(col_idx)
                        break

                if not duration_col_idx:
                    print(f"      ⚠️  Skipping - No DEVICE KERNEL DURATION column found")
                    continue

                # Check if formatted column already exists
                formatted_exists = False
                for col_idx, cell in enumerate(ws[1], 1):
                    if cell.value and (
                        "DURATION FORMATTED" in str(cell.value) or "DURATION [ns] - FORMATTED" in str(cell.value)
                    ):
                        formatted_exists = True
                        break

                if not formatted_exists:
                    # Insert formatted column after Column C (CORE COUNT) - position 4
                    new_col_idx = 4  # Column D (after A, B, C)

                    # Insert a new column at position 4, shifting everything right
                    ws.insert_cols(new_col_idx)

                    # Update duration column reference if it was shifted by the insertion
                    if duration_col_idx >= new_col_idx:
                        duration_col_idx += 1
                        duration_col_letter = get_column_letter(duration_col_idx)

                    new_col_letter = get_column_letter(new_col_idx)

                    # Add header for new column
                    ws[f"{new_col_letter}1"] = "DURATION [ns] - FORMATTED"
                    ws[f"{new_col_letter}1"].font = Font(bold=True)

                    # Count rows with data
                    max_row = ws.max_row
                    data_rows = 0

                    # Process data rows (starting from row 2)
                    for row_idx in range(2, max_row + 1):
                        original_cell = ws[f"{duration_col_letter}{row_idx}"]
                        new_cell = ws[f"{new_col_letter}{row_idx}"]

                        if original_cell.value is not None:
                            try:
                                # Convert to number (remove commas if present)
                                if isinstance(original_cell.value, str):
                                    # Remove commas and convert to int
                                    numeric_value = int(original_cell.value.replace(",", ""))
                                else:
                                    numeric_value = int(original_cell.value)

                                new_cell.value = numeric_value
                                new_cell.style = comma_style
                                data_rows += 1

                            except (ValueError, AttributeError):
                                # If conversion fails, copy as-is
                                new_cell.value = original_cell.value

                    if data_rows > 0:
                        # Apply Data Bar conditional formatting
                        # Define the range for data bars (excluding header)
                        data_range = f"{new_col_letter}2:{new_col_letter}{max_row}"

                        # Create data bar rule with blue gradient
                        data_bar_rule = DataBarRule(
                            start_type="min",
                            start_value=None,
                            end_type="max",
                            end_value=None,
                            color="5B9BD5",  # Blue color
                            showValue=True,
                            minLength=None,
                            maxLength=None,
                        )

                        # Apply the rule to the range
                        ws.conditional_formatting.add(data_range, data_bar_rule)

                        print(f"      ✅ Added formatted column (Column D) with {data_rows} data bars")
                        processed_sheets += 1
                    else:
                        print(f"      ⚠️  No data rows found")
                else:
                    print(f"      ℹ️  Formatted column already exists")
                    processed_sheets += 1

            except Exception as e:
                print(f"      ❌ Error processing sheet: {e}")
                continue

        # Save the modified workbook
        wb.save(self.excel_file)
        print(f"✅ Data bars added to {processed_sheets} sheets")

        return wb

    def add_fps_calculations(self):
        """Add FPS calculations to all worksheets"""
        print(f"\n🏃‍♂️ Adding FPS calculations...")

        # Load workbook
        wb = openpyxl.load_workbook(self.excel_file)

        processed_sheets = 0
        fps_results = []

        for sheet_idx, ws in enumerate(wb.worksheets, 1):
            sheet_name = ws.title
            print(f"  [{sheet_idx:2d}/{len(wb.worksheets)}] Processing sheet: '{sheet_name}'")

            try:
                # Find the formatted duration column
                formatted_col_idx = None
                formatted_col_letter = None

                # Check first row for headers
                for col_idx, cell in enumerate(ws[1], 1):
                    if cell.value and (
                        "DURATION FORMATTED" in str(cell.value) or "DURATION [ns] - FORMATTED" in str(cell.value)
                    ):
                        formatted_col_idx = col_idx
                        formatted_col_letter = get_column_letter(col_idx)
                        break

                if not formatted_col_idx:
                    print(f"      ⚠️  Skipping - No formatted duration column found")
                    continue

                # Find last row with data in the formatted column
                max_row = ws.max_row
                last_data_row = 1
                total_duration_ns = 0
                valid_values = 0

                # Sum all values in the formatted duration column
                for row_idx in range(2, max_row + 1):
                    cell = ws[f"{formatted_col_letter}{row_idx}"]
                    if cell.value is not None:
                        try:
                            # Convert to number
                            if isinstance(cell.value, str):
                                numeric_value = float(cell.value.replace(",", ""))
                            else:
                                numeric_value = float(cell.value)

                            total_duration_ns += numeric_value
                            valid_values += 1
                            last_data_row = row_idx

                        except (ValueError, AttributeError):
                            continue

                if total_duration_ns > 0:
                    # Calculate FPS: FPS = 10^9 / total_duration_ns
                    fps = (10**9) / total_duration_ns

                    # Check if FPS calculation already exists
                    fps_exists = False
                    for row_idx in range(last_data_row + 1, last_data_row + 5):
                        cell = ws[f"A{row_idx}"]
                        if cell.value and "FPS" in str(cell.value):
                            fps_exists = True
                            break

                    if not fps_exists:
                        # Add some spacing and then the FPS calculation
                        fps_row = last_data_row + 2

                        # Add "TOTAL DURATION (ns):" label and value
                        ws[f"A{fps_row}"] = "TOTAL DURATION (ns):"
                        ws[f"A{fps_row}"].font = Font(bold=True)
                        ws[f"{formatted_col_letter}{fps_row}"] = total_duration_ns
                        ws[f"{formatted_col_letter}{fps_row}"].number_format = "#,##0"

                        # Add "FPS (Frames/Second):" label and value
                        fps_row += 1
                        ws[f"A{fps_row}"] = "FPS (Frames/Second):"
                        ws[f"A{fps_row}"].font = Font(bold=True, color="0000FF")  # Blue text
                        ws[f"{formatted_col_letter}{fps_row}"] = fps
                        ws[f"{formatted_col_letter}{fps_row}"].number_format = "0.00"
                        ws[f"{formatted_col_letter}{fps_row}"].font = Font(bold=True, color="0000FF")
                        ws[f"{formatted_col_letter}{fps_row}"].fill = PatternFill(
                            start_color="E6F3FF", end_color="E6F3FF", fill_type="solid"
                        )

                    # Store results for summary
                    fps_results.append(
                        {
                            "Model": sheet_name,
                            "Total_Duration_ns": total_duration_ns,
                            "Total_Duration_ms": total_duration_ns / 1_000_000,
                            "FPS": fps,
                            "Operations": valid_values,
                        }
                    )

                    print(f"      ✅ Total: {total_duration_ns:,.0f} ns | FPS: {fps:.2f}")
                    processed_sheets += 1
                else:
                    print(f"      ⚠️  No valid duration data found")

            except Exception as e:
                print(f"      ❌ Error processing sheet: {e}")
                continue

        # Save the modified workbook
        wb.save(self.excel_file)
        print(f"✅ FPS calculations added to {processed_sheets} sheets")

        return fps_results

    def print_summary(self, fps_results):
        """Print final summary of processing results"""
        print("\n" + "=" * 80)
        print("🎉 EXCEL PROCESSING COMPLETE")
        print("=" * 80)
        print(f"📁 Output file: {self.excel_file}")
        print(f"✅ Features added:")
        print(f"   • Opcodes column with cleaned operation names")
        print(f"   • Formatted duration column (Column D) with comma separators")
        print(f"   • Blue gradient data bars for visual comparison")
        print(f"   • FPS calculations for performance analysis")

        # Display FPS summary sorted by FPS (fastest first)
        if fps_results:
            fps_results.sort(key=lambda x: x["FPS"], reverse=True)

            print(f"\n🏆 MODEL PERFORMANCE RANKING (by FPS):")
            print("-" * 80)
            print(f"{'Rank':<4} {'Model':<25} {'FPS':<8} {'Duration (ms)':<15} {'Ops':<5}")
            print("-" * 80)

            for rank, result in enumerate(fps_results, 1):
                model_name = result["Model"][:23] + "..." if len(result["Model"]) > 23 else result["Model"]
                print(
                    f"{rank:<4} {model_name:<25} {result['FPS']:<8.2f} {result['Total_Duration_ms']:<15.2f} {result['Operations']:<5}"
                )

            if len(fps_results) > 1:
                print(f"\n📊 PERFORMANCE INSIGHTS:")
                print(f"🥇 Fastest Model: {fps_results[0]['Model']} ({fps_results[0]['FPS']:.2f} FPS)")
                print(f"🐌 Slowest Model: {fps_results[-1]['Model']} ({fps_results[-1]['FPS']:.2f} FPS)")
                print(f"⚡ Speed Difference: {fps_results[0]['FPS']/fps_results[-1]['FPS']:.1f}x faster")

    def process(self):
        """Main processing method"""
        print("🚀 Starting Excel Processing Pipeline")
        print("=" * 80)

        # Step 1: Handle input file
        if self.input_file.suffix.lower() == ".csv":
            self.excel_file = self.process_csv_to_excel()
        elif self.input_file.suffix.lower() == ".xlsx":
            self.excel_file = self.load_excel_file()
        else:
            raise ValueError(f"Unsupported file format: {self.input_file.suffix}")

        # Step 2: Add formatted columns and data bars
        self.add_formatted_columns_and_databars()

        # Step 3: Add FPS calculations
        fps_results = self.add_fps_calculations()

        # Step 4: Print summary
        self.print_summary(fps_results)


def main():
    """Main entry point with argument parsing"""
    parser = argparse.ArgumentParser(
        description="Combined Excel/CSV Processor for Performance Analysis",
        epilog="""
Examples:
  python excel_processor_combined.py data.csv
  python excel_processor_combined.py data.xlsx ./output/
  python excel_processor_combined.py performance_data.csv /path/to/output/
        """,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )

    parser.add_argument("input_file", help="Input file (CSV or Excel)")
    parser.add_argument("output_dir", nargs="?", default=".", help="Output directory (default: current directory)")

    args = parser.parse_args()

    # Check if input file exists
    if not Path(args.input_file).exists():
        print(f"❌ Error: Input file '{args.input_file}' not found!")
        return 1

    try:
        # Create processor and run
        processor = ExcelProcessor(args.input_file, args.output_dir)
        processor.process()
        return 0

    except Exception as e:
        print(f"❌ Error: {e}")
        return 1


if __name__ == "__main__":
    sys.exit(main())
