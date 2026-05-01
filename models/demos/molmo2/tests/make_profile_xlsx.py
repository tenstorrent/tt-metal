# SPDX-FileCopyrightText: © 2025 Tenstorrent USA, Inc.
# SPDX-License-Identifier: Apache-2.0
"""Build a multi-sheet Excel workbook from per-block Tracy ops-perf CSV files.

Usage (after running run_block_profiles.sh):
  python models/demos/molmo2/tests/make_profile_xlsx.py \
      --csv-dir generated/profiler/block_csvs \
      --output molmo2_block_profile.xlsx

Each CSV file corresponds to one block (named <block_name>_ops.csv).
One Excel sheet is created per block, filtered to device 0, containing:
  - Op Code, Math Fidelity, Core Count, Parallelization Strategy
  - Device Kernel Duration [us]
  - Input 0..3: shape (WxZxYxX), Layout, Dtype, Memory / Sharding
  - Output 0..3: same
  - Attributes (truncated), Program Cache Hit

Colour-coding:
  - Matmul rows: light blue
  - AllBroadcast / CCL rows: light orange (bottleneck indicator)
  - SDPA rows: light green
"""

import argparse
import csv
import pathlib
import re

# openpyxl is bundled with most Python envs; install with: pip install openpyxl
try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    raise SystemExit("openpyxl not installed — run: pip install openpyxl")

# ── Column spec ────────────────────────────────────────────────────────────
KEEP_COLS = [
    "OP CODE",
    "MATH FIDELITY",
    "CORE COUNT",
    "PARALLELIZATION STRATEGY",
    "DEVICE KERNEL DURATION [ns]",  # converted to µs in output
    "DEVICE KERNEL DURATION PER CORE MIN [ns]",
    "DEVICE KERNEL DURATION PER CORE MAX [ns]",
]
N_TENSORS = 4  # capture up to 4 inputs and 4 outputs

TENSOR_FIELDS = ["W_PAD[LOGICAL]", "Z_PAD[LOGICAL]", "Y_PAD[LOGICAL]", "X_PAD[LOGICAL]", "LAYOUT", "DATATYPE", "MEMORY"]

MISC_COLS = [
    "PROGRAM CACHE HIT",
    "ATTRIBUTES",
]

# Row colours (ARGB hex)
FILL_HEADER = PatternFill("solid", fgColor="1F4E79")  # dark blue header
FILL_MATMUL = PatternFill("solid", fgColor="DDEEFF")  # light blue
FILL_CCL = PatternFill("solid", fgColor="FFE0B2")  # light orange
FILL_SDPA = PatternFill("solid", fgColor="DDFFDD")  # light green
FILL_NORM = PatternFill("solid", fgColor="F5F5DC")  # beige
FILL_ALT = PatternFill("solid", fgColor="F7F7F7")  # very light grey

HEADER_FONT = Font(color="FFFFFF", bold=True)


def _shape(row, prefix, idx):
    """Return 'WxZxYxX' string, or '' if all empty."""
    parts = []
    for dim in ("W_PAD[LOGICAL]", "Z_PAD[LOGICAL]", "Y_PAD[LOGICAL]", "X_PAD[LOGICAL]"):
        key = f"{prefix}_{idx}_{dim}"
        v = row.get(key, "").strip()
        # strip the '[N]' padded form to show only logical value, e.g. '4096[4096]' → '4096'
        if v:
            m = re.match(r"(\d+)\[(\d+)\]", v)
            v = m.group(2) if m else v
        parts.append(v)
    if not any(parts):
        return ""
    return "×".join(p if p else "?" for p in parts)


def _mem_short(mem_str):
    """Shorten the verbose MemoryConfig string to something readable."""
    if not mem_str:
        return ""
    mem_str = mem_str.strip()
    if "HEIGHT_SHARDED" in mem_str:
        return "L1_H_SHARD"
    if "WIDTH_SHARDED" in mem_str:
        return "L1_W_SHARD"
    if "BLOCK_SHARDED" in mem_str:
        return "L1_BLK_SHARD"
    if "L1" in mem_str and "INTERLEAVED" in mem_str:
        return "L1_INTERLEAVED"
    if "DRAM" in mem_str and "INTERLEAVED" in mem_str:
        return "DRAM_INTERLEAVED"
    if "DRAM" in mem_str:
        return "DRAM"
    if "L1" in mem_str:
        return "L1"
    return mem_str[:30]


def _attr_short(attr_str):
    """Trim the verbose attributes string."""
    if not attr_str:
        return ""
    # remove outer braces, truncate
    s = attr_str.strip().lstrip("{").rstrip("}")
    return s[:120] + ("…" if len(s) > 120 else "")


def build_header_row():
    row = [
        "Op Code",
        "Math Fidelity",
        "Cores",
        "Parallelization",
        "Kernel [µs]",
        "Kernel/Core Min [µs]",
        "Kernel/Core Max [µs]",
    ]
    for i in range(N_TENSORS):
        row += [f"In{i} Shape", f"In{i} Layout", f"In{i} Dtype", f"In{i} Memory"]
    for i in range(N_TENSORS):
        row += [f"Out{i} Shape", f"Out{i} Layout", f"Out{i} Dtype", f"Out{i} Memory"]
    row += ["Cache Hit", "Attributes"]
    return row


def csv_row_to_excel(row):
    def ns_to_us(v):
        try:
            return round(float(v) / 1000, 3) if v.strip() else ""
        except (ValueError, AttributeError):
            return ""

    out = [
        row.get("OP CODE", ""),
        row.get("MATH FIDELITY", ""),
        row.get("CORE COUNT", ""),
        row.get("PARALLELIZATION STRATEGY", ""),
        ns_to_us(row.get("DEVICE KERNEL DURATION [ns]", "")),
        ns_to_us(row.get("DEVICE KERNEL DURATION PER CORE MIN [ns]", "")),
        ns_to_us(row.get("DEVICE KERNEL DURATION PER CORE MAX [ns]", "")),
    ]
    for i in range(N_TENSORS):
        shape = _shape(row, "INPUT", i)
        layout = row.get(f"INPUT_{i}_LAYOUT", "").strip()
        dtype = row.get(f"INPUT_{i}_DATATYPE", "").strip()
        mem = _mem_short(row.get(f"INPUT_{i}_MEMORY", ""))
        out += [shape, layout, dtype, mem]
    for i in range(N_TENSORS):
        shape = _shape(row, "OUTPUT", i)
        layout = row.get(f"OUTPUT_{i}_LAYOUT", "").strip()
        dtype = row.get(f"OUTPUT_{i}_DATATYPE", "").strip()
        mem = _mem_short(row.get(f"OUTPUT_{i}_MEMORY", ""))
        out += [shape, layout, dtype, mem]
    out += [
        row.get("PROGRAM CACHE HIT", ""),
        _attr_short(row.get("ATTRIBUTES", "")),
    ]
    return out


def row_fill(op_code):
    oc = op_code.lower()
    if "matmul" in oc:
        return FILL_MATMUL
    if "allbroadcast" in oc or "allgather" in oc or "reducescatter" in oc or "ccl" in oc:
        return FILL_CCL
    if "sdpa" in oc:
        return FILL_SDPA
    if "layernorm" in oc or "rmsnorm" in oc or "moreh" in oc:
        return FILL_NORM
    return None


def add_sheet(wb, sheet_name, csv_path):
    ws = wb.create_sheet(title=sheet_name)

    with open(csv_path, newline="") as f:
        reader = csv.DictReader(f)
        all_rows = [r for r in reader if r.get("DEVICE ID", "").strip() == "0"]

    if not all_rows:
        ws.append(["(no device-0 ops found)"])
        return

    # Header
    header = build_header_row()
    ws.append(header)
    for col_idx, title in enumerate(header, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = FILL_HEADER
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    # Data rows
    for row_idx, csv_r in enumerate(all_rows, 2):
        data = csv_row_to_excel(csv_r)
        ws.append(data)
        fill = row_fill(csv_r.get("OP CODE", ""))
        alt = FILL_ALT if row_idx % 2 == 0 else None
        applied = fill or alt
        if applied:
            for col_idx in range(1, len(header) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = applied

    # Freeze header, auto-filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Column widths
    col_widths = {
        1: 38,  # Op Code
        2: 12,  # Math Fidelity
        3: 7,  # Cores
        4: 18,  # Parallelization
        5: 12,  # Kernel [µs]
        6: 16,  # Kernel/Core Min
        7: 16,  # Kernel/Core Max
    }
    # Tensor columns: 4 per tensor × (4 in + 4 out) = 32 cols starting at 8
    for i, w in enumerate([16, 8, 14, 18] * (N_TENSORS * 2)):
        col_widths[8 + i] = w
    # Last 2 cols
    col_widths[8 + N_TENSORS * 2 * 4] = 10  # Cache Hit
    col_widths[8 + N_TENSORS * 2 * 4 + 1] = 50  # Attributes

    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Summary row at bottom
    ws.append([])
    total_us = sum(float(r.get("DEVICE KERNEL DURATION [ns]", 0) or 0) / 1000 for r in all_rows)
    ws.append([f"TOTAL ops: {len(all_rows)}", "", "", "TOTAL kernel [µs]:", round(total_us, 1)])


def add_summary_sheet(wb, block_stats):
    """First sheet: per-block summary table."""
    ws = wb.create_sheet(title="Summary", index=0)
    header = [
        "Block",
        "Ops (dev0)",
        "Total Kernel [µs]",
        "Matmul [µs]",
        "AllBroadcast/CCL [µs]",
        "SDPA [µs]",
        "Other [µs]",
    ]
    ws.append(header)
    for col_idx, title in enumerate(header, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = FILL_HEADER
        cell.font = HEADER_FONT

    for block, stats in block_stats.items():
        ws.append(
            [
                block,
                stats["n_ops"],
                round(stats["total_us"], 1),
                round(stats["matmul_us"], 1),
                round(stats["ccl_us"], 1),
                round(stats["sdpa_us"], 1),
                round(stats["other_us"], 1),
            ]
        )

    for col_idx in range(1, len(header) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 22
    ws.freeze_panes = "A2"


def collect_stats(csv_path):
    with open(csv_path, newline="") as f:
        reader = csv.DictReader(f)
        rows = [r for r in reader if r.get("DEVICE ID", "").strip() == "0"]

    stats = {"n_ops": len(rows), "total_us": 0, "matmul_us": 0, "ccl_us": 0, "sdpa_us": 0, "other_us": 0}
    for r in rows:
        dur_us = float(r.get("DEVICE KERNEL DURATION [ns]", 0) or 0) / 1000
        oc = r.get("OP CODE", "").lower()
        stats["total_us"] += dur_us
        if "matmul" in oc:
            stats["matmul_us"] += dur_us
        elif "allbroadcast" in oc or "allgather" in oc or "reducescatter" in oc:
            stats["ccl_us"] += dur_us
        elif "sdpa" in oc:
            stats["sdpa_us"] += dur_us
        else:
            stats["other_us"] += dur_us
    return stats


def main():
    ap = argparse.ArgumentParser(description="Build per-block profile Excel from Tracy CSVs.")
    ap.add_argument(
        "--csv-dir", default="generated/profiler/block_csvs", help="Directory containing <block>_ops.csv files"
    )
    ap.add_argument("--output", default="molmo2_block_profile.xlsx", help="Output .xlsx path")
    ap.add_argument("--csv", nargs="*", help="Explicit CSV files to include (overrides --csv-dir)")
    args = ap.parse_args()

    # Collect CSVs
    if args.csv:
        csv_files = [(pathlib.Path(f).stem.replace("_ops", ""), pathlib.Path(f)) for f in args.csv]
    else:
        csv_dir = pathlib.Path(args.csv_dir)
        csv_files = sorted((p.stem.replace("_ops", ""), p) for p in csv_dir.glob("*_ops.csv"))

    if not csv_files:
        raise SystemExit(f"No CSV files found. Provide --csv-dir or --csv.")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    block_stats = {}
    for block_name, csv_path in csv_files:
        print(f"  Adding sheet: {block_name}  ({csv_path})")
        block_stats[block_name] = collect_stats(csv_path)
        add_sheet(wb, block_name, csv_path)

    add_summary_sheet(wb, block_stats)

    out_path = pathlib.Path(args.output)
    wb.save(out_path)
    print(f"\nSaved: {out_path}  ({out_path.stat().st_size // 1024} KB)")
    print(f"Sheets: Summary + {[n for n, _ in csv_files]}")


if __name__ == "__main__":
    main()
