#!/usr/bin/env python3
"""
Combined Excel/CSV Processor
=============================
This script combines the functionality of:
1. CSV processing with opcodes extraction (perf_analyzer.py)
2. Excel formatting with data bars (format_excel_databars.py)
3. FPS calculations (add_fps_calculation.py)

Usage:
    python excel_processor_combined.py <input_path> [output_directory]

    input_path: Can be either:
                - A single .csv or .xlsx file
                - A folder containing .csv files (searches recursively in subdirectories)
                  Creates combined Excel with tabs for each CSV found
                  Note: Automatically ignores profile_log_device.csv and per_core_op_to_op_times*.csv files
    output_directory: Optional, defaults to current directory
"""

import sys
import os
import pandas as pd
import openpyxl
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import NamedStyle, Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import argparse
from pathlib import Path
import glob


class ExcelProcessor:
    def __init__(self, input_path, output_dir="."):
        self.input_path = Path(input_path)
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        self.workbook = None
        self.excel_file = None
        self.csv_files = []
        self.is_folder_mode = False

    def _clean_sheet_name(self, name):
        """Clean sheet name for Excel compatibility"""
        # Remove invalid Excel sheet name characters
        invalid_chars = ["[", "]", ":", "*", "?", "/", "\\"]
        for char in invalid_chars:
            name = name.replace(char, "")

        # Limit to 31 characters (Excel limit)
        return name[:31] if len(name) > 31 else name

    def _ensure_unique_sheet_name(self, sheet_name, existing_sheets):
        """Ensure sheet name is unique by appending counter if needed"""
        original_name = sheet_name
        counter = 1
        existing_names = [s["name"] for s in existing_sheets]

        while sheet_name in existing_names:
            # Keep room for counter suffix
            base_name = original_name[:28] if len(original_name) > 28 else original_name
            sheet_name = f"{base_name}_{counter}"
            counter += 1

        return sheet_name

    def detect_input_type(self):
        """Detect if input is a file or folder and prepare CSV file list"""
        if self.input_path.is_dir():
            print(f"📁 Folder mode detected: {self.input_path}")
            self.is_folder_mode = True

            # Find all CSV files recursively in the folder and subfolders
            csv_pattern = str(self.input_path / "**" / "*.csv")
            all_csv_files = [Path(f) for f in glob.glob(csv_pattern, recursive=True)]

            # Filter out profile_log_device.csv and per_core_op_to_op_times*.csv files
            self.csv_files = [
                f
                for f in all_csv_files
                if f.name != "profile_log_device.csv" and not f.name.startswith("per_core_op_to_op_times")
            ]

            # Report filtering if any files were excluded
            excluded_count = len(all_csv_files) - len(self.csv_files)
            if excluded_count > 0:
                excluded_types = []
                profile_log_count = len([f for f in all_csv_files if f.name == "profile_log_device.csv"])
                per_core_count = len([f for f in all_csv_files if f.name.startswith("per_core_op_to_op_times")])

                if profile_log_count > 0:
                    excluded_types.append(f"{profile_log_count} profile_log_device.csv file(s)")
                if per_core_count > 0:
                    excluded_types.append(f"{per_core_count} per_core_op_to_op_times*.csv file(s)")

                print(f"📋 Filtered out {excluded_count} files: {', '.join(excluded_types)}")

            if not self.csv_files:
                if excluded_count > 0:
                    raise ValueError(
                        f"No processable CSV files found in folder (excluding {excluded_count} system files): {self.input_path}"
                    )
                else:
                    raise ValueError(f"No CSV files found in folder (including subfolders): {self.input_path}")

            # Sort files for consistent processing order
            self.csv_files.sort()
            print(f"📊 Found {len(self.csv_files)} CSV files (including subfolders):")
            for i, csv_file in enumerate(self.csv_files, 1):
                # Show relative path from input directory for better readability
                relative_path = csv_file.relative_to(self.input_path)
                print(f"  [{i:2d}] {relative_path}")

        elif self.input_path.is_file():
            if self.input_path.suffix.lower() == ".csv":
                print(f"📄 Single CSV file detected: {self.input_path}")
                self.csv_files = [self.input_path]
                self.is_folder_mode = False
            elif self.input_path.suffix.lower() == ".xlsx":
                print(f"📊 Excel file detected: {self.input_path}")
                self.csv_files = []
                self.is_folder_mode = False
            else:
                raise ValueError(f"Unsupported file format: {self.input_path.suffix}")
        else:
            raise ValueError(f"Input path not found: {self.input_path}")

    def process_single_csv(self, csv_file, sheet_name=None):
        """Process a single CSV file and return the processed DataFrame"""
        print(f"📄 Processing CSV file: {csv_file}")

        try:
            # Read CSV with better error handling
            try:
                df = pd.read_csv(csv_file)
            except pd.errors.EmptyDataError:
                print(f"      ❌ Error: CSV file is empty: {csv_file}")
                return None
            except pd.errors.ParserError as e:
                print(f"      ❌ Error: Failed to parse CSV file {csv_file}: {e}")
                return None
            except UnicodeDecodeError as e:
                print(f"      ❌ Error: Encoding issue in CSV file {csv_file}: {e}")
                try:
                    # Try with different encoding
                    df = pd.read_csv(csv_file, encoding="latin-1")
                    print(f"      ✅ Successfully read with latin-1 encoding")
                except Exception as e2:
                    print(f"      ❌ Error: Failed with alternative encoding: {e2}")
                    return None

            # Check if DataFrame is empty
            if df.empty:
                print(f"      ⚠️  Warning: CSV file contains no data: {csv_file}")
                return None

            print(f"      📊 Raw data: {len(df)} rows, {len(df.columns)} columns")

            # Define required columns, with fallback options for different CSV formats
            required_cols = ["OP CODE", "CORE COUNT", "DEVICE KERNEL DURATION [ns]"]

            # Check if required columns exist
            missing_cols = [col for col in required_cols if col not in df.columns]
            if missing_cols:
                print(f"      ⚠️  Warning: Missing required columns {missing_cols} in {csv_file}")
                print(f"      📋 Available columns: {list(df.columns)}")

                # Try to find similar columns
                if "OP CODE" not in df.columns:
                    # Look for similar column names
                    op_cols = [col for col in df.columns if "OP" in col.upper() and "CODE" in col.upper()]
                    if op_cols:
                        print(f"      💡 Found similar OP CODE column: {op_cols[0]}")
                        df.rename(columns={op_cols[0]: "OP CODE"}, inplace=True)
                    else:
                        print(f"      ❌ Cannot process file - no OP CODE column found")
                        return None

                # Check for CORE COUNT alternatives
                if "CORE COUNT" not in df.columns:
                    core_cols = [col for col in df.columns if "CORE" in col.upper() and "COUNT" in col.upper()]
                    if core_cols:
                        print(f"      💡 Found similar CORE COUNT column: {core_cols[0]}")
                        df.rename(columns={core_cols[0]: "CORE COUNT"}, inplace=True)
                    else:
                        print(f"      ❌ Cannot process file - no CORE COUNT column found")
                        return None

                # Check for DURATION alternatives
                if "DEVICE KERNEL DURATION [ns]" not in df.columns:
                    duration_cols = [col for col in df.columns if "DURATION" in col.upper() and "ns" in col.lower()]
                    if not duration_cols:
                        duration_cols = [col for col in df.columns if "DURATION" in col.upper()]
                    if duration_cols:
                        print(f"      💡 Found similar DURATION column: {duration_cols[0]}")
                        df.rename(columns={duration_cols[0]: "DEVICE KERNEL DURATION [ns]"}, inplace=True)
                    else:
                        print(f"      ❌ Cannot process file - no DURATION column found")
                        return None

                # Check again after renaming
                missing_cols = [col for col in required_cols if col not in df.columns]
                if missing_cols:
                    print(f"      ❌ Cannot process file - still missing: {missing_cols}")
                    return None

            # Add MATH FIDELITY if available
            if "MATH FIDELITY" in df.columns:
                required_cols.append("MATH FIDELITY")

            optional_cols = []

            # Check which input/output columns exist and add them
            for prefix in ["INPUT_0", "OUTPUT_0"]:
                for suffix in ["W", "Z", "Y", "X"]:
                    # Try different column name formats
                    possible_names = [
                        f"{prefix}_{suffix}",  # Original format
                        f"{prefix}_{suffix}_PAD[LOGICAL]",  # New format with _PAD[LOGICAL]
                    ]
                    for col_name in possible_names:
                        if col_name in df.columns:
                            optional_cols.append(col_name)
                            break

                # Add layout, datatype, and memory columns
                for attr in ["LAYOUT", "DATATYPE", "MEMORY"]:
                    col_name = f"{prefix}_{attr}"
                    if col_name in df.columns:
                        optional_cols.append(col_name)

            # Select available columns
            col_list = required_cols + optional_cols
            available_cols = [col for col in col_list if col in df.columns]

            print(f"      📊 Processing {len(available_cols)} available columns")
            df = df[available_cols]

            # Find first non-nan row in CORE COUNT with error handling
            try:
                first_index = df["CORE COUNT"].first_valid_index()
                if first_index is not None:
                    df = df.iloc[first_index:]
                else:
                    print(f"      ⚠️  Warning: No valid CORE COUNT data found")
            except Exception as e:
                print(f"      ⚠️  Warning: Error processing CORE COUNT column: {e}")

            # Check if DataFrame is empty after filtering
            if df.empty:
                print(f"      ⚠️  Warning: No valid data after filtering: {csv_file}")
                return None

            # Add OP ID as first column (row numbers starting from 1)
            df.reset_index(drop=True, inplace=True)
            df.insert(0, "OP ID", range(1, len(df) + 1))

            # Rename columns for brevity
            df.rename(columns=lambda x: x.replace("INPUT_", "IN_"), inplace=True)
            df.rename(columns=lambda x: x.replace("OUTPUT_", "OUT_"), inplace=True)

            # Clean up OP CODE values with error handling
            try:
                df = df.replace("InterleavedToShardedDeviceOperation", "I2S", regex=True)
                df = df.replace("ShardedToInterleavedDeviceOperation", "S2I", regex=True)
                df = df.replace("DeviceOperation", "", regex=True)
                df = df.replace("DEV_0_", "", regex=True)
            except Exception as e:
                print(f"      ⚠️  Warning: Error cleaning OP CODE values: {e}")

            # Convert data types with error handling
            try:
                df = df.convert_dtypes()
            except Exception as e:
                print(f"      ⚠️  Warning: Error converting data types: {e}")

            # Calculate total duration before formatting with error handling
            try:
                total_duration = df["DEVICE KERNEL DURATION [ns]"].sum()
                print(f"      📈 Total duration: {total_duration:,} ns ({len(df)} operations)")
            except Exception as e:
                print(f"      ⚠️  Warning: Error calculating total duration: {e}")
                print(f"      📈 Processed {len(df)} operations")

            return df

        except FileNotFoundError:
            print(f"      ❌ Error: File not found: {csv_file}")
            return None
        except PermissionError:
            print(f"      ❌ Error: Permission denied accessing file: {csv_file}")
            return None
        except Exception as e:
            print(f"      ❌ Error: Unexpected error processing CSV file {csv_file}: {type(e).__name__}: {e}")
            return None

    def process_csv_to_excel(self):
        """Process CSV file(s) and create Excel with opcodes and formatted columns"""
        if self.csv_files:
            # Multiple CSV files or single CSV file
            return self.process_multiple_csvs_to_excel()
        else:
            # Single Excel file (backwards compatibility)
            return self.load_excel_file()

    def process_multiple_csvs_to_excel(self):
        """Process multiple CSV files and create a combined Excel with separate sheets"""
        print(f"📊 Processing {len(self.csv_files)} CSV files into combined Excel...")

        # Generate output filename
        if self.is_folder_mode:
            base_name = self.input_path.name + "_combined"
        else:
            base_name = self.csv_files[0].stem + "_processed"

        self.excel_file = self.output_dir / f"{base_name}_complete.xlsx"

        # Track successful sheets
        successful_sheets = []
        failed_files = []

        # Create Excel writer
        with pd.ExcelWriter(self.excel_file, engine="openpyxl") as writer:
            for i, csv_file in enumerate(self.csv_files, 1):
                print(f"  [{i:2d}/{len(self.csv_files)}] Processing: {csv_file.name}")

                try:
                    # Process the CSV file
                    df = self.process_single_csv(csv_file)

                    # Skip if DataFrame is None or empty
                    if df is None or df.empty:
                        print(f"      ⚠️  Skipping - no valid data")
                        failed_files.append((csv_file.name, "No valid data"))
                        continue

                    # Create clean sheet name from CSV filename
                    sheet_name = self._clean_sheet_name(csv_file.stem)

                    # Ensure unique sheet name
                    sheet_name = self._ensure_unique_sheet_name(sheet_name, successful_sheets)

                    # Write to Excel sheet
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    successful_sheets.append({"name": sheet_name, "file": csv_file.name, "rows": len(df)})
                    print(f"      ✅ Created sheet: '{sheet_name}' with {len(df)} rows")

                except Exception as e:
                    print(f"      ❌ Error processing {csv_file.name}: {type(e).__name__}: {e}")
                    failed_files.append((csv_file.name, f"Processing error: {e}"))
                    continue

        # Check if any sheets were created successfully
        if not successful_sheets:
            raise ValueError(
                f"No valid sheets could be created from the CSV files. All {len(self.csv_files)} files failed processing."
            )

        # Report results
        print(f"✅ Combined Excel file created: {self.excel_file}")
        print(f"   📊 Successfully processed: {len(successful_sheets)}/{len(self.csv_files)} files")

        if failed_files:
            print(f"   ⚠️  Failed files: {len(failed_files)}")
            for failed_file, error_reason in failed_files:
                print(f"      - {failed_file}: {error_reason}")

        print(
            f"   📊 Success rate: {len(successful_sheets)}/{len(self.csv_files)} ({len(successful_sheets)/len(self.csv_files)*100:.1f}%)"
        )

        return self.excel_file

    def load_excel_file(self):
        """Load existing Excel file"""
        print(f"📊 Loading Excel file: {self.input_path}")
        self.excel_file = self.input_path
        return self.excel_file

    def add_formatted_columns_and_databars(self):
        """Add formatted duration column with data bars to all worksheets"""
        print(f"\n🎨 Adding formatted columns and data bars...")

        # Load workbook
        wb = openpyxl.load_workbook(self.excel_file)

        # Create number style with comma separator
        comma_style = NamedStyle(name="comma_no_decimal")
        comma_style.number_format = "#,##0"

        try:
            wb.add_named_style(comma_style)
        except ValueError:
            # Style already exists
            pass

        processed_sheets = 0

        for sheet_idx, ws in enumerate(wb.worksheets, 1):
            sheet_name = ws.title
            print(f"  [{sheet_idx:2d}/{len(wb.worksheets)}] Processing sheet: '{sheet_name}'")

            try:
                # Find the DEVICE KERNEL DURATION column
                duration_col_idx = None
                duration_col_letter = None

                # Check first row for headers
                for col_idx, cell in enumerate(ws[1], 1):
                    if cell.value and "DEVICE KERNEL DURATION [ns]" in str(cell.value):
                        duration_col_idx = col_idx
                        duration_col_letter = get_column_letter(col_idx)
                        break

                if not duration_col_idx:
                    print(f"      ⚠️  Skipping - No DEVICE KERNEL DURATION column found")
                    continue

                # Check if formatted column already exists
                formatted_exists = False
                for col_idx, cell in enumerate(ws[1], 1):
                    if cell.value and (
                        "DURATION FORMATTED" in str(cell.value) or "DURATION [ns] - FORMATTED" in str(cell.value)
                    ):
                        formatted_exists = True
                        break

                if not formatted_exists:
                    # Insert formatted column after Column C (CORE COUNT) - position 4
                    new_col_idx = 4  # Column D (after A, B, C)

                    # Insert a new column at position 4, shifting everything right
                    ws.insert_cols(new_col_idx)

                    # Update duration column reference if it was shifted by the insertion
                    if duration_col_idx >= new_col_idx:
                        duration_col_idx += 1
                        duration_col_letter = get_column_letter(duration_col_idx)

                    new_col_letter = get_column_letter(new_col_idx)

                    # Add header for new column
                    ws[f"{new_col_letter}1"] = "DURATION [ns] - FORMATTED"
                    ws[f"{new_col_letter}1"].font = Font(bold=True)

                    # Count rows with data
                    max_row = ws.max_row
                    data_rows = 0

                    # Process data rows (starting from row 2)
                    for row_idx in range(2, max_row + 1):
                        try:
                            original_cell = ws[f"{duration_col_letter}{row_idx}"]
                            new_cell = ws[f"{new_col_letter}{row_idx}"]

                            if original_cell.value is not None:
                                try:
                                    # Convert to number (remove commas if present)
                                    if isinstance(original_cell.value, str):
                                        # Remove commas and convert to int
                                        numeric_value = int(original_cell.value.replace(",", ""))
                                    else:
                                        numeric_value = int(original_cell.value)

                                    new_cell.value = numeric_value
                                    new_cell.style = comma_style
                                    data_rows += 1

                                except (ValueError, AttributeError, TypeError) as e:
                                    # If conversion fails, copy as-is
                                    new_cell.value = original_cell.value
                                    print(f"      ⚠️  Warning: Could not convert value in row {row_idx}: {e}")
                        except Exception as e:
                            print(f"      ⚠️  Warning: Error processing row {row_idx}: {e}")
                            continue

                    if data_rows > 0:
                        try:
                            # Apply Data Bar conditional formatting
                            # Define the range for data bars (excluding header)
                            data_range = f"{new_col_letter}2:{new_col_letter}{max_row}"

                            # Create data bar rule with blue gradient
                            data_bar_rule = DataBarRule(
                                start_type="min",
                                start_value=None,
                                end_type="max",
                                end_value=None,
                                color="5B9BD5",  # Blue color
                                showValue=True,
                                minLength=None,
                                maxLength=None,
                            )

                            # Apply the rule to the range
                            ws.conditional_formatting.add(data_range, data_bar_rule)

                            print(f"      ✅ Added formatted column (Column D) with {data_rows} data bars")
                            processed_sheets += 1
                        except Exception as e:
                            print(f"      ⚠️  Warning: Could not add data bars: {e}")
                            processed_sheets += 1
                    else:
                        print(f"      ⚠️  No data rows found")
                else:
                    print(f"      ℹ️  Formatted column already exists")
                    processed_sheets += 1

            except Exception as e:
                print(f"      ❌ Error processing sheet '{sheet_name}': {type(e).__name__}: {e}")
                continue

        # Save the modified workbook
        wb.save(self.excel_file)
        print(f"✅ Data bars added to {processed_sheets} sheets")

        return wb

    def add_fps_calculations(self):
        """Add FPS calculations to all worksheets"""
        print(f"\n🏃‍♂️ Adding FPS calculations...")

        # Load workbook
        wb = openpyxl.load_workbook(self.excel_file)

        processed_sheets = 0
        fps_results = []

        for sheet_idx, ws in enumerate(wb.worksheets, 1):
            sheet_name = ws.title
            print(f"  [{sheet_idx:2d}/{len(wb.worksheets)}] Processing sheet: '{sheet_name}'")

            try:
                # Find the formatted duration column
                formatted_col_idx = None
                formatted_col_letter = None

                # Check first row for headers
                for col_idx, cell in enumerate(ws[1], 1):
                    if cell.value and (
                        "DURATION FORMATTED" in str(cell.value) or "DURATION [ns] - FORMATTED" in str(cell.value)
                    ):
                        formatted_col_idx = col_idx
                        formatted_col_letter = get_column_letter(col_idx)
                        break

                if not formatted_col_idx:
                    print(f"      ⚠️  Skipping - No formatted duration column found")
                    continue

                # Find last row with data in the formatted column
                max_row = ws.max_row
                last_data_row = 1
                total_duration_ns = 0
                valid_values = 0

                # Sum all values in the formatted duration column
                for row_idx in range(2, max_row + 1):
                    cell = ws[f"{formatted_col_letter}{row_idx}"]
                    if cell.value is not None:
                        try:
                            # Convert to number
                            if isinstance(cell.value, str):
                                numeric_value = float(cell.value.replace(",", ""))
                            else:
                                numeric_value = float(cell.value)

                            total_duration_ns += numeric_value
                            valid_values += 1
                            last_data_row = row_idx

                        except (ValueError, AttributeError):
                            continue

                if total_duration_ns > 0:
                    # Calculate FPS: FPS = 10^9 / total_duration_ns
                    fps = (10**9) / total_duration_ns

                    # Check if FPS calculation already exists
                    fps_exists = False
                    for row_idx in range(last_data_row + 1, last_data_row + 5):
                        cell = ws[f"A{row_idx}"]
                        if cell.value and "FPS" in str(cell.value):
                            fps_exists = True
                            break

                    if not fps_exists:
                        # Add some spacing and then the FPS calculation
                        fps_row = last_data_row + 2

                        # Add "TOTAL DURATION (ns):" label and value
                        ws[f"A{fps_row}"] = "TOTAL DURATION (ns):"
                        ws[f"A{fps_row}"].font = Font(bold=True)
                        ws[f"{formatted_col_letter}{fps_row}"] = total_duration_ns
                        ws[f"{formatted_col_letter}{fps_row}"].number_format = "#,##0"

                        # Add "FPS (Frames/Second):" label and value
                        fps_row += 1
                        ws[f"A{fps_row}"] = "FPS (Frames/Second):"
                        ws[f"A{fps_row}"].font = Font(bold=True, color="0000FF")  # Blue text
                        ws[f"{formatted_col_letter}{fps_row}"] = fps
                        ws[f"{formatted_col_letter}{fps_row}"].number_format = "0.00"
                        ws[f"{formatted_col_letter}{fps_row}"].font = Font(bold=True, color="0000FF")
                        ws[f"{formatted_col_letter}{fps_row}"].fill = PatternFill(
                            start_color="E6F3FF", end_color="E6F3FF", fill_type="solid"
                        )

                    # Store results for summary
                    fps_results.append(
                        {
                            "Model": sheet_name,
                            "Total_Duration_ns": total_duration_ns,
                            "Total_Duration_ms": total_duration_ns / 1_000_000,
                            "FPS": fps,
                            "Operations": valid_values,
                        }
                    )

                    print(f"      ✅ Total: {total_duration_ns:,.0f} ns | FPS: {fps:.2f}")
                    processed_sheets += 1
                else:
                    print(f"      ⚠️  No valid duration data found")

            except Exception as e:
                print(f"      ❌ Error processing sheet: {e}")
                continue

        # Save the modified workbook
        wb.save(self.excel_file)
        print(f"✅ FPS calculations added to {processed_sheets} sheets")

        return fps_results

    def print_summary(self, fps_results):
        """Print final summary of processing results"""
        print("\n" + "=" * 80)
        print("🎉 EXCEL PROCESSING COMPLETE")
        print("=" * 80)
        print(f"📁 Output file: {self.excel_file}")

        if self.is_folder_mode:
            print(f"📊 Input: {len(self.csv_files)} CSV files from folder '{self.input_path.name}'")
        else:
            print(f"📊 Input: Single file processing")

        print(f"✅ Features added:")
        print(f"   • Opcodes column with cleaned operation names")
        print(f"   • Formatted duration column (Column D) with comma separators")
        print(f"   • Blue gradient data bars for visual comparison")
        print(f"   • FPS calculations for performance analysis")

        if self.is_folder_mode and len(self.csv_files) > 1:
            print(f"   • Combined Excel with {len(self.csv_files)} separate sheets")

        print(f"\n🎯 All data processed and formatted successfully!")
        if self.is_folder_mode:
            print(f"💡 Each CSV file is now a separate sheet in the Excel workbook.")

    def process(self):
        """Main processing method"""
        print("🚀 Starting Excel Processing Pipeline")
        print("=" * 80)

        # Step 1: Detect input type and prepare file list
        self.detect_input_type()

        # Step 2: Handle input file(s)
        if self.csv_files:
            # Process CSV files (single or multiple)
            self.excel_file = self.process_csv_to_excel()
        elif self.input_path.suffix.lower() == ".xlsx":
            # Load existing Excel file
            self.excel_file = self.load_excel_file()
        else:
            raise ValueError(f"No valid input files found")

        # Step 3: Add formatted columns and data bars
        self.add_formatted_columns_and_databars()

        # Step 4: Add FPS calculations
        fps_results = self.add_fps_calculations()

        # Step 5: Print summary
        self.print_summary(fps_results)


def main():
    """Main entry point with argument parsing"""
    parser = argparse.ArgumentParser(
        description="Combined Excel/CSV Processor for Performance Analysis",
        epilog="""
Examples:
  python excel_processor_combined.py data.csv
  python excel_processor_combined.py data.xlsx ./output/
  python excel_processor_combined.py /path/to/csv/folder/           # Searches recursively
  python excel_processor_combined.py /path/to/csv/folder/ ./output/ # Recursive with output dir
        """,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )

    parser.add_argument(
        "input_path", help="Input path (CSV file, Excel file, or folder to search recursively for CSV files)"
    )
    parser.add_argument("output_dir", nargs="?", default=".", help="Output directory (default: current directory)")

    args = parser.parse_args()

    # Check if input path exists
    if not Path(args.input_path).exists():
        print(f"❌ Error: Input path '{args.input_path}' not found!")
        return 1

    try:
        # Create processor and run
        processor = ExcelProcessor(args.input_path, args.output_dir)
        processor.process()
        return 0

    except Exception as e:
        print(f"❌ Error: {e}")
        return 1


if __name__ == "__main__":
    sys.exit(main())
